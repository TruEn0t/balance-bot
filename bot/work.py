from PIL import Image, ImageDraw, ImageFont
import openpyxl

def can_float(num):
    try:
        float(num.replace(',', '.'))
        return True

    except:
        return False


def can_int(num):return num.isdigit()


def find_data(text):
    txt = ''

    num = 0

    for word in text.split(' '):
        if can_float(word.replace('+', '').replace('-', '')):
            num = float(word.replace(',', '.'))
            num = -num if not '+' in word else num

            if '-' in word:
                num = -num

        else:
            txt += ' ' + word


    if txt != '' and num:
        return [txt, num]

    else:
        return None


def create_photo(document_name):
    wb = openpyxl.load_workbook(document_name)
    ws = wb[wb.sheetnames[0]]

    img = Image.new('RGB', (500,1000 if ws.max_row > 50 else 20 * ws.max_row+20),(255,255,255))
    font = ImageFont.truetype('Roboto-Black.ttf', size=13)
    draw = ImageDraw.Draw(img)


    draw.rectangle((0, 0, 500, 1000), (255, 255, 255), (187, 187, 187))
    for x in range(50):
        draw.line((0, (x+1)*20, 500, (x+1)*20), (187, 187, 187), 1)

    for x in range(4):
        draw.line((x*125, 0, x*125, 1000),(187, 187, 187),1)

    row = 0

    sum = 0
    pls = 0
    mns = 0

    for x in ws:
        if x[0].row < ws.max_row - 50:continue

        date = x[0].value
        comment = x[1].value
        m = x[2].value
        p = x[3].value

        if date != 'Дата':
            pls += p if p else 0
            mns += m if m else 0

        d = 0
        for x in [date, comment, m, p]:
            if x == None:x = ''
            draw.text((125 * d + 5, 20 * row+6), str(x), fill=(0,0,0),font=font)

            d += 1

        row += 1

    sum = pls - mns

    draw.text((125*1,20*row+1),str(sum),fill=(0,0,0),font = font)
    draw.text((125 * 2 + 5, 20 * row + 1), str(mns), fill=(0, 0, 0), font=font)
    draw.text((125 * 3 + 5, 20 * row + 1), str(pls), fill=(0, 0, 0), font=font)

    img.save('temp.png')

    return 'temp.png'

