import telebot
import json
from work import *
import datetime
from openpyxl import Workbook, load_workbook
import os

config = json.load(open('config.json'))

bot = telebot.TeleBot(config['token'])

@bot.message_handler(content_types=['text'])
def new_message(message):

    data = find_data(message.text)
    xls_path = str(message.chat.id) + '.xls'

    bot.send_message(message.chat.id, data.__str__())

    if not os.path.isfile(xls_path):
    	wb = Workbook()
    	wb.save(xls_path)

    wb = load_workbook(xls_path)

bot.polling(none_stop=True, interval=0)